import os
import datetime
from openpyxl import load_workbook
from openpyxl import Workbook

st_name = 'Anoop'


# Get current date
today_date = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M")
def mark_present(st_name):
    names = os.listdir('output/')
    # print(names)


    sub = 'Attendance_' + today_date

    # Check if the attendance file for the current date exists
    attendance_file = 'attendance/' + sub + '.xlsx'
    if not os.path.exists(attendance_file):
        workbook = Workbook()
        print("Creating Spreadsheet with Title: " + sub)
        sheet = workbook.active
        sheet.title = sub  # Set sheet title to the current date
        count = 2
        for i in names:
            sheet.cell(row=count, column=1).value = i
            count += 1
        workbook.save(attendance_file)

    # Load workbook using openpyxl
    wb = load_workbook(attendance_file)

    # Get reference to active sheet
    sheet = wb.active

    # Write current timestamp to cell B2
    sheet['B2'] = str(datetime.datetime.now())

    count = 2
    for i in names:
        if i in st_name:
            sheet.cell(row=count, column=2).value = 'P'
        else:
            sheet.cell(row=count, column=2).value = 'A'
        sheet.cell(row=count, column=1).value = i
        count += 1

    wb.save(attendance_file)


mark_present(st_name)

# import os.path
# import datetime
# from google.auth.transport.requests import Request
# from google.oauth2.credentials import Credentials
# from google_auth_oauthlib.flow import InstalledAppFlow
# from googleapiclient.discovery import build
# from googleapiclient.errors import HttpError

# # If modifying these scopes, delete the file token.json.
# SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

# # The ID of your Google Spreadsheet.
# SPREADSHEET_ID = "18_Ypfyq5vlGlHmrB00D_wcDKu9V9N981IbE3qelRa0w"

# st_name = 'Anoop'


# def mark_present(st_name):
#     global SPREADSHEET_ID
#     creds = None
#     try:
#         if os.path.exists("token.json"):
#             creds = Credentials.from_authorized_user_file("token.json", SCOPES)
#     except FileNotFoundError as e:
#         print(f"Error: {e}. Please ensure 'token.json' file exists.")
#         return

#     if not creds or not creds.valid:
#         if creds and creds.expired and creds.refresh_token:
#             creds.refresh(Request())
#         else:
#             try:
#                 flow = InstalledAppFlow.from_client_secrets_file(
#                     "credentials.json", SCOPES)
#                 creds = flow.run_local_server(port=0)
#                 with open("token.json", "w") as token:
#                     token.write(creds.to_json())
#             except FileNotFoundError as e:
#                 print(
#                     f"Error: {e}. Please ensure 'credentials.json' file exists.")
#                 return
#             except Exception as e:
#                 print(f"Error: {e}. Failed to get credentials.")
#                 return

#     try:
#         service = build("sheets", "v4", credentials=creds)
#         # Fetch names from the "output" folder
#         try:
#             names = os.listdir('output/')
#             num_students = len(names)
#             last_row = num_students + 1
#         except OSError as e:
#             print(f"Error: {e}. Failed to access 'output' folder.")
#             return

#         # Get current date
#         today_date = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M")
#         sub = 'Attendance_' + today_date
        
#         # Check if the attendance file for the current date exists
#         range_name = f"Sheet1!A1:C1"
#         result = service.spreadsheets().values().get(
#             spreadsheetId=SPREADSHEET_ID, range=range_name).execute()
#         values = result.get("values", [])

#         # If the spreadsheet doesn't exist, create a new one
#         if not values:
#             print("Creating new spreadsheet...")
#             spreadsheet = {
#                 'properties': {
#                     'title': sub
#                 }
#             }
#             spreadsheet = service.spreadsheets().create(body=spreadsheet,
#                                                         fields='spreadsheetId').execute()
#             SPREADSHEET_ID = spreadsheet.get('spreadsheetId')
#             # Write header labels to the new spreadsheet
#             header_labels = ["Timestamp", "Roll No", "Status"]
#             header_values = [header_labels]
#             header_body = {'values': header_values}
#             service.spreadsheets().values().update(spreadsheetId=SPREADSHEET_ID, range=range_name, valueInputOption='RAW', body=header_body).execute()

#         # Construct the range for the header row
#         header_range = f"Sheet1!A1:C1"

#         # Define the header labels
#         header_labels = ["Timestamp", "Roll No", "Status"]

#         # Write the header labels to the spreadsheet
#         header_values = [header_labels]
#         body = {'values': header_values}
#         service.spreadsheets().values().update(spreadsheetId=SPREADSHEET_ID, range=header_range, valueInputOption='RAW', body=body).execute()

#         # Get the number of students
#         num_students = len(names)

#         # Determine the last row of the range
#         last_row = num_students + 1  # Adding 1 to account for the header row

#         # Construct the range dynamically
#         range_name = f"Sheet1!A{last_row}:C{last_row+1}"

#         # Write current timestamp and mark attendance for each student
#         now = str(datetime.datetime.now().strftime("%H-%M"))
#         for name in names:
#             if name in st_name:
#                 attendance = 'P'
#             else:
#                 attendance = 'A'
#             values.append([now, name, attendance])  # Assuming 'P' for present

#         # Update values in the spreadsheet
#         body = {'values': values}
#         service.spreadsheets().values().update(spreadsheetId=SPREADSHEET_ID,
#                                                range=range_name, valueInputOption='RAW', body=body).execute()

#     except HttpError as err:
#         print(f"Google Sheets API Error: {err}")


# mark_present(st_name)

